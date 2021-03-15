"""
create update excel
2020-03-07
"""
import os
import time

from openpyxl import Workbook, load_workbook

excel_name = '预警通报爬取统计.xlsx'
worksheet_name = '预警通报爬取统计'


def init_excel():
    row_1 = ['预警时间', '预警通报名称', '通报来源', '通报链接', '简述', '重点漏洞或漏洞详情', '修复建议']
    workbook = Workbook()
    global worksheet_name
    worksheet = workbook.create_sheet(worksheet_name)
    worksheet.append(row_1)
    global excel_name
    workbook.save(excel_name)


def to_timestamp(str_date):
    # 转换成时间数组
    time_array = time.strptime(str_date, "%Y-%m-%d %H:%M:%S")
    # 返回时间戳
    return time.mktime(time_array)


def sort_by_time(notices):
    count = len(notices)
    for i in range(1, count):
        key = notices[i]
        j = i - 1
        while j >= 0:
            if to_timestamp(notices[j][0]) < to_timestamp(key[0]):
                notices[j+1] = notices[j]
                notices[j] = key
            j -= 1
    return notices


def excel_add(notices):
    global excel_name
    if not os.path.exists(excel_name):
        init_excel()
    workbook = load_workbook(excel_name)
    global worksheet_name
    worksheet = workbook[worksheet_name]
    # 获取excel 公告 url
    col_c = []
    for i in worksheet['D']:
        col_c.append(i.value)
    # 比较url是否相同，不相同的录入excel
    # 根据时间排序
    for i in sort_by_time(notices):
        if i[3] not in col_c:
            worksheet.append(i)
    workbook.save(excel_name)
