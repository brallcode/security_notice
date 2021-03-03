import os
import re
import json
import time

from bs4 import BeautifulSoup
import urllib
import jsonpath
from openpyxl import Workbook
from openpyxl import load_workbook
import requests

headers = {
    'Host': 'cert.360.cn',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:85.0) Gecko/20100101 Firefox/85.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate',
}


# 360cert爬取网页
def get_text(tmp_url):
    add2 = urllib.request.Request(url=tmp_url, headers=headers)
    html2 = urllib.request.urlopen(url=tmp_url, timeout=10)
    soup = BeautifulSoup(html2.read(), features="html.parser")
    # 获取标题
    title = soup.find("div", {"class": "detail-title-head"})
    notice_name = title.get_text()
    # print(title.get_text())

    # 获取事件描述
    notice_description = ''
    # event_description = soup.find("div", {"class": "news-content"})
    # print(event_description.get_text())
    """
    # 遍历通报链接
    for i in url_infos:
        add2 = urllib.request.Request(url=i, headers=headers)
        html2 = urllib.request.urlopen(url=i, timeout=10)
        print(html2.read())
    """

    # 第一个h2 到 第二个 h2之间为事件描述
    for i in soup.find("h2").next_siblings:
        if i.name == "h2":
            break
        # print(i.get_text())
        notice_description = notice_description + i.get_text()
    # 获取重点漏洞 || 漏洞详情
    info = ''
    for i in soup.findAll("h2"):
        # 查找重点漏洞tag
        tag_zhongd = re.match(".*重点漏洞", i.string)
        # 获取终端漏洞描述
        if tag_zhongd is not None:
            for j in i.next_siblings:
                if j.name == "h2":
                    break
                info = info + j.get_text()
            break
        # print(i.get_text())
        # 查找漏洞详情
        tag_loudongxq = re.match(".*漏洞详情", i.string)
        if tag_loudongxq is not None:
            for j in i.next_siblings:
                if j.name == "h2":
                    break
                info = info + j.get_text()
    # 获取修复建议
    h2_jy = ''
    for i in soup.findAll("h2"):
        h2_jy_name = re.match(".*修复建议", i.string)
        if h2_jy_name is None:
            continue
        for j in i.next_siblings:
            if j.name == 'h2':
                break
            h2_jy = h2_jy + j.get_text()
    """    
    # 分别获取通用、临时修复建议
    for i in soup.findAll("h2"):
        h2_jy = re.match(".*修复建议", i.string)
        if h2_jy is None:
            continue
        # print(i.next_sibling)
        # 通用修复建议
        h3_ty = i.next_sibling
        for i in h3_ty.next_siblings:
            if i.name == "h2":
                # 获取 临时建议 tag
                text_ty = re.match("临时修复建议", i.string)
                if text_ty is None:
                    h3_ls = None
                else:
                    h3_ls = i
                break
            # print(i.get_text())
        # 临时修复建议
        # if h3_ls is not None:
    """

    # get release date: release_date
    release_date_tag = soup.find("div", {"class": "detail-news-date"})
    release_date_match = re.match("\d{4}-\d{1,2}-\d{1,2}", release_date_tag.string)
    release_date = release_date_match.group()

    # paqu xinxi list
    notice = [release_date, notice_name, tmp_url, notice_description, info, h2_jy]

    # print(notice)
    return notice


# init excel
def excel_init(name):
    workbook = Workbook()
    row_1 = ['预警时间', '预警通报名称', '通报链接', '简述', '重点漏洞或漏洞详情', '修复建议']
    # worksheet = workbook.active
    ws1 = workbook.create_sheet('sheet1')
    ws1.append(row_1)
    workbook.save(name)


def main_360cert():
    url = "https://cert.360.cn/warning/searchbypage?length=6&start=0"

    # html = requests.get(url=url, headers=headers)
    add = urllib.request.Request(url=url, headers=headers)
    html = urllib.request.urlopen(url=url, timeout=10)
    # print(html.read())
    html_json = json.loads(html.read())
    # print(html_json)
    # 获取id值
    id_list = jsonpath.jsonpath(html_json, "$..id")
    # 获取title
    title_list = jsonpath.jsonpath(html_json, '$..title')
    # print(title_list)
    # print(id_list)
    # 排除 安全事件周报
    id_list_del = []
    for i in range(0, len(title_list)):
        if re.match('安全事件周报.*', title_list[i]):
            id_list_del.append(id_list[i])

    for i in id_list_del:
        id_list.remove(i)

    # print(id_list)
    # print(id_list)
    # id_json = json.loads(html_json['data'])
    # print(html_json['data'][0]['id'])

    url_prefix = "https://cert.360.cn/warning/detail?id="
    url_infos = []
    for i in id_list:
        url_infos.append(url_prefix + i)

    #    print(url_prefix + i)
    # print(url_infos)

    notices = []
    for i in url_infos:
        notices.append(get_text(i))
    # print(len(notices))

    excel_name = '360cert预警通报统计.xlsx'
    if not os.path.exists(excel_name):
        excel_init(excel_name)
    workbook = load_workbook(excel_name)
    worksheet = workbook['sheet1']
    max_row = worksheet.max_row
    for i in range(2, worksheet.max_row + 1):
        print(worksheet.cell(row=i, column=2).value)

    # 提取已经存在通告
    index_esists = []
    for i in worksheet.rows:
        # print(i[1].value)
        index_esists.append(i[1].value)
    # 判断是否重复，未则写入
    for i in notices:
        if i[1] not in index_esists:
            worksheet.append(i)
    # worksheet.append(notice)
    workbook.save(excel_name)


# main_360cert()

def split_one_enter(str1):
    return str1.split('\n', maxsplit=1)[1]


def get_nox_detail(id_detail):
    headers_detail = {
        'Host': 'nox.qianxin.com',
        'User-Agent': 'Mozilla/5.0(X11; Linux x86_64; rv: 85.0) Gecko/20100101 Firefox/85.0',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-US,en; q = 0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Content-Type': 'application/json; charset=utf-8',
        'Content-Length': '12',
        'Origin': 'https://nox.qianxin.com',
        'Connection': 'close',
        'Referer': 'https://nox.qianxin.com/article/' + str(id_detail),
        'Cache-Control': 'max-age=0'
    }
    # print(headers_detail['Referer'])
    # data
    data_detail = {'id': str(id_detail)}
    # url
    url_detail = 'https://nox.qianxin.com/api/web/portal/article/info/show'
    reponse_detail = requests.post(url_detail, headers=headers_detail, data=json.dumps(data_detail))
    # print(reponse_detail.text)
    # format to json
    reponse_detail_json = json.loads(reponse_detail.text)
    # get data json
    detail_data = jsonpath.jsonpath(reponse_detail_json, '$[data]')[0]
    detail_list = detail_data['article_content'].split('##')
    # get brief introduction || security notice
    # 删除开头‘安全通告’
    brief_introduction = split_one_enter(detail_list[1])
    # 获取漏洞描述
    vulnerability_des = split_one_enter(detail_list[2])
    # 风险等级
    risk_level = split_one_enter(detail_list[3])
    # 影响范围
    sphere_of_influence = split_one_enter(detail_list[4])
    # 处置建议
    disposal_recom = split_one_enter(detail_list[5])
    # 访问链接
    access_link = 'https://nox.qianxin.com/article/' + str(id_detail)
    # 最近更新时间
    last_update_time = detail_data['last_update_time']
    # 转换成时间戳
    time_array = time.strptime(last_update_time, "%Y-%m-%d %H:%M:%S")
    timestamp = time.mktime(time_array)
    # 标题
    title = detail_data['title']
    print(timestamp)


def main_nox():
    # nox安全监测平台（奇安信cert）
    # 首页通告信息列表接口
    url = 'https://nox.qianxin.com/api/web/portal/article/info/list'
    # response = urllib.request.urlopen(url=url, headers=headers)
    # print(response.read())
    # 构造cookie
    t = int(time.time())
    cookie = "Hm_lvt_9ecbd54546eb318ba9fbdfb7dfee53da={}; Hm_lpvt_9ecbd54546eb318ba9fbdfb7dfee53da={}".format(t, t)
    # HTTP header
    # header list
    headers_list = {
        'Host': 'nox.qianxin.com',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:85.0) Gecko/20100101 Firefox/85.0',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Content-Type': 'application/json; charset=utf-8',
        'Content-Length': '54',
        'Origin': 'https://nox.qianxin.com',
        'Connection': 'close',
        'Referer': 'https://nox.qianxin.com/risk',
        'Cache-Control': 'max-age=0'
    }
    # data list
    data_list = {"page_no": 1, "page_size": 10, "category": "风险通告"}
    reponse_list = requests.post(url, headers=headers_list, data=json.dumps(data_list))
    # 将将响应文本转换为json格式
    data_list_json = json.loads(reponse_list.text)
    # print(data_list_json)
    # 获取id值，用于拼接访问通告详情的url
    id_list = jsonpath.jsonpath(data_list_json, "$..id")
    # 获取通告title
    title_list = jsonpath.jsonpath(data_list_json, '$..title')
    # print(id_list)
    # print(title_list)

    # test get_nox_detail()
    get_nox_detail(id_list[0])


main_nox()

"""
腾讯威胁通告: https://s.tencent.com/research/bsafe
通告信息在每个div标签 class="newsli"中
"""


def main_tencent():
    headers = {
        "Host": "s.tencent.com",
        "User-Agent": "Mozilla/5.0(X11; Linux x86_64; rv:85.0) Gecko/20100101 Firefox/85.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "close",
        "Upgrade-Insecure-Requests": "1"
    }

    # 绿盟威胁通告： http://blog.nsfocus.net/category/threat-alert
    # 启明安全通告： https://www.venustech.com.cn/new_type/aqtg/
    # 深信服漏洞预警： https://sec.sangfor.com.cn/wiki-safe-events
